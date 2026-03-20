import io
import re
from typing import Dict, List, Any
from openpyxl import load_workbook, Workbook
from django.db import transaction
from django.contrib.auth.models import User
from paasify.models import UserProfile, Subject, UserProject

class ExcelImporterService:
    @staticmethod
    def generate_template(template_type: str = 'professor') -> bytes:
        """
        Genera un archivo Excel en memoria con las cabeceras requeridas.
        """
        wb = Workbook()
        ws = wb.active
        
        if template_type == 'professor':
            ws.title = "Alumnos y Proyectos"
            # Cabeceras requeridas para el profesor
            ws.append(["Nombre de usuario", "Nombre", "Apellidos", "Email", "Contraseña", "Nombre Proyecto"])
            # Ejemplo
            ws.append(["alumno1", "Alumno", "Uno", "alumno1@urjc.es", "pass1234", "TFG_Alumno1"])
            ws.append(["alumno2", "Alumno", "Dos", "alumno2@urjc.es", "", ""])
            
            # Auto-ajustar ancho
            for col in range(1, 7):
                ws.column_dimensions[chr(64 + col)].width = 25
        else:
            ws.title = "Usuarios Admin"
            # Cabeceras requeridas para el django admin
            ws.append(["Nombre de usuario", "Nombre", "Apellidos", "Email", "Contraseña", "Rol"])
            ws.append(["admin1", "Admin", "Principal", "admin1@urjc.es", "securepass", "admin"])
            ws.append(["profesor1", "Profesor", "Uno", "profesor1@urjc.es", "", "teacher"])
            ws.append(["alumno1", "Alumno", "Uno", "alumno1@urjc.es", "", "student"])
            
            for col in range(1, 7):
                ws.column_dimensions[chr(64 + col)].width = 25
                
        # Cabeceras en negrita
        for cell in ws[1]:
            cell.font = cell.font.copy(bold=True)
            
        output = io.BytesIO()
        wb.save(output) # Guardamos en Memoria
        return output.getvalue()

    @staticmethod
    def _is_valid_email(email: str) -> bool:
        regex = r'^\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Z|a-z]{2,7}\b'
        return re.match(regex, email) is not None

    @staticmethod
    def process_professor_import(file_obj, subject: Subject, confirm: bool = False) -> Dict[str, Any]:
        """
        Procesa un archivo Excel para la subida desde el portal de Profesor.
        confirm=False: Lanza Preview (No guarda nada).
        confirm=True: Guarda la informacion en BD.
        """
        try:
            wb = load_workbook(filename=file_obj, read_only=True, data_only=True)
            if len(wb.sheetnames) > 1:
                return {"error": "El archivo Excel contiene múltiples hojas. Solo se admite un archivo con una única hoja.", "rows": []}
            ws = wb.active
        except Exception as e:
            return {"error": "Formato de archivo inválido. Asegúrese de subir un .xlsx.", "rows": []}

        rows_result = []
        has_errors = False
        
        # Validar la cabecera (primera linea vacia o incompleta)
        header = [str(cell.value).strip().lower() if cell.value else "" for cell in ws[1]]
        if 'email' not in header or 'nombre de usuario' not in header:
             return {"error": "El excel debe contener al menos las columnas 'Nombre de usuario' y 'Email'. Por favor, descargue la plantilla.", "rows": []}

        # Índices de columnas
        idx_username = header.index('nombre de usuario')
        idx_name = header.index('nombre') if 'nombre' in header else -1
        idx_lastname = header.index('apellidos') if 'apellidos' in header else -1
        idx_email = header.index('email')
        idx_password = header.index('contraseña') if 'contraseña' in header else -1
        idx_project = header.index('nombre proyecto') if 'nombre proyecto' in header else -1

        all_usernames_in_sheet = set()
        all_emails_in_sheet = set()
        all_projects_in_sheet = set()

        processed_data = []

        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            # Si la fila entera está vacía, saltar
            if not any(row):
                continue
                
            username = str(row[idx_username] or "").strip() if 0 <= idx_username < len(row) else ""
            name = str(row[idx_name] or "").strip() if 0 <= idx_name < len(row) else ""
            lastname = str(row[idx_lastname] or "").strip() if 0 <= idx_lastname < len(row) else ""
            email = str(row[idx_email] or "").strip().lower() if 0 <= idx_email < len(row) else ""
            password = str(row[idx_password] or "").strip() if 0 <= idx_password < len(row) else ""
            project_name = str(row[idx_project] or "").strip() if 0 <= idx_project < len(row) else ""
            
            row_status = "ok" # 'ok', 'warning', 'error'
            messages = []
            
            # --- VALIDACIONES BÁSICAS ---
            if not username:
                row_status = "error"
                messages.append("El campo Nombre de usuario es obligatorio.")
            else:
                if username in all_usernames_in_sheet:
                    row_status = "error"
                    messages.append("Nombre de usuario duplicado en el propio Excel.")
                all_usernames_in_sheet.add(username)

            if not email:
                row_status = "error"
                messages.append("El campo Email es obligatorio.")
            elif not ExcelImporterService._is_valid_email(email):
                row_status = "error"
                messages.append(f"El Email '{email}' tiene un formato inválido.")
            else:
                if email in all_emails_in_sheet:
                    row_status = "error"
                    messages.append("Email duplicado en el propio Excel.")
                all_emails_in_sheet.add(email)

            if project_name:
                if project_name in all_projects_in_sheet:
                    row_status = "error"
                    messages.append("Nombre de proyecto duplicado en el propio Excel.")
                else:
                    all_projects_in_sheet.add(project_name)
                    
                    # Validar si el proyecto ya existe en la asignatura (globalmente o por nombre)
                    # En la BD real el UserProject no fuerza unique global de name, pero es buena práctica no duplicar nombres en la misma Asignatura por claridad
                    if UserProject.objects.filter(subject=subject, place=project_name).exists():
                        row_status = "error"
                        messages.append(f"El proyecto '{project_name}' ya existe en esta asignatura.")

            # --- VERIFICACIÓN EN BASE DE DATOS ---
            user_exists = False
            user_already_in_subject = False
            project_is_new = False # Flag para UI
            
            if row_status != "error":
                # Check username or email matching an existing user
                from django.db.models import Q
                user_obj = User.objects.filter(Q(username=username) | Q(email=email)).first()
                if user_obj:
                    user_exists = True
                    # Validate that both match if the user already exists, or give a warning
                    if user_obj.username != username or user_obj.email != email:
                        row_status = "error"
                        messages.append(f"Conflicto de identidad: ya existe en DB (User: {user_obj.username}, Email: {user_obj.email}).")
                        messages.append("Sugerencia: Use los datos exactos que ya existen en PaaSify para vincularlo, o elija un nombre de usuario y email que no estén en uso.")
                    
                    try:
                        if user_obj in subject.students.all():
                            user_already_in_subject = True
                    except Exception:
                        pass
                
                if user_exists and row_status != "error":
                    if not user_already_in_subject:
                        # Alumno existe pero no en la asig
                        if project_name:
                            messages.append("Usuario ya existe en PaaSify. Se le vinculará a la asignatura y se creará el proyecto.")
                        else:
                            messages.append("Usuario ya existe en PaaSify. Simplemente se le matriculará en la asignatura.")
                            project_is_new = False
                        row_status = "warning"
                    else: # user_already_in_subject is True
                        if project_name:
                            # Check if project already exists in this subject
                            if UserProject.objects.filter(subject=subject, place=project_name).exists():
                                messages.append(f"El proyecto '{project_name}' ya existe en esta asignatura.")
                                row_status = "error" # Project name conflict for existing user in subject
                            else:
                                messages.append(f"El usuario ya está matriculado en la asignatura. Se le creará el nuevo proyecto.")
                                row_status = "matriculado"
                                project_is_new = True
                        else:
                            row_status = "matriculado"
                            messages.append("El usuario ya está matriculado en esta asignatura y no se pide crear proyecto. Fila ignorada.")
                            project_is_new = False
                elif not user_exists:
                    if project_name:
                        messages.append("Usuario nuevo. Se creará su cuenta, se le matriculará y se generará su proyecto.")
                        project_is_new = True
                    else:
                        messages.append("Usuario nuevo. Solo se creará la cuenta y se le matriculará.")
                        project_is_new = False
                # If row_status is already error, project_is_new should remain False (default)
                # If row_status is 'ok' and project_name exists, it's a new project for a new user.
                # This is handled in the 'elif not user_exists' block.
                # If row_status is 'ok' and project_name does not exist, project_is_new is False.
                # If row_status is 'warning' (user exists, not in subject), project_is_new is set there.
                # If row_status is 'matriculado' (user exists, in subject), project_is_new is set there.
                # So, no general assignment for project_is_new is needed here.
            
            if row_status == 'error':
                has_errors = True

            # Registrar data curada
            processed_data.append({
                "row": row_num,
                "username": username,
                "name": name,
                "lastname": lastname,
                "email": email,
                "password": password,
                "project_name": project_name,
                "project_is_new": project_is_new,
                "user_exists": user_exists,
                "status": row_status,
                "messages": messages
            })
            
        result_payload = {
            "error": None,
            "has_errors": has_errors,
            "total_rows": len(processed_data),
            "rows": processed_data
        }

        # Si NO es confirmar, devolver preview
        if not confirm:
            return result_payload
            
        # Si es confirmar y hay errores, no dejamos pasar.
        if confirm and has_errors:
            return {"error": "Existen filas con errores. Corríjalas antes de confirmar.", "rows": processed_data}
            
        # EJECUCIÓN 
        if confirm and not has_errors:
            try:
                with transaction.atomic():
                    for data in processed_data:
                        username = data['username']
                        email = data['email']
                        pwd = data['password'] or email # Fallback a email de pass si no se proveyó
                        project_name = data['project_name']
                        
                        user = User.objects.filter(username=username).first()
                        if not user:
                            # CREAR NUEVO
                            user = User.objects.create_user(
                                username=username,
                                email=email,
                                password=pwd,
                                first_name=data['name'],
                                last_name=data['lastname']
                            )
                            # Perfil Alumno automatico
                            from paasify.roles import ensure_user_group, STUDENT_GROUP_NAMES, DEFAULT_STUDENT_GROUP
                            ensure_user_group(user, STUDENT_GROUP_NAMES, DEFAULT_STUDENT_GROUP)
                            profile, _ = UserProfile.objects.update_or_create(
                                user=user,
                                defaults={
                                    'nombre': f"{data['name']} {data['lastname']}".strip() or data['username'],
                                    'year': email,
                                    'must_change_password': not bool(data['password'])
                                }
                            )
                            # Matricular
                            subject.students.add(user)
                        else:
                            # VINCULAR EXISTENTE
                            subject.students.add(user)
                            
                        # Mapear Proyecto
                        if project_name:
                            UserProject.objects.create(
                                place=project_name, 
                                user_profile=user.user_profile,
                                subject=subject
                            )
                            
            except Exception as e:
                return {"error": f"Error fatal procesando base de datos: {str(e)}", "rows": processed_data}

        # Todo ejecutado bien
        return {"error": None, "has_errors": False, "success": True, "rows": processed_data}

    @staticmethod
    def process_admin_import(file_obj) -> Dict[str, Any]:
        """
        Procesa el import estricto y bloqueante de Django Admin.
        No hay preview, es "todo o nada".
        Retorna success=True o raise Exception interceptable por messages.error.
        """
        try:
            wb = load_workbook(filename=file_obj, read_only=True, data_only=True)
            if len(wb.sheetnames) > 1:
                raise Exception("El archivo Excel contiene múltiples hojas. Solo se admite un archivo con una única hoja.")
            ws = wb.active
        except Exception as e:
            if "múltiples hojas" in str(e):
                raise
            raise Exception("Formato de archivo inválido. Asegúrese de subir un .xlsx.")

        header = [str(cell.value).strip().lower() if cell.value else "" for cell in ws[1]]
        if 'email' not in header or 'nombre de usuario' not in header:
            raise Exception("El excel debe contener al menos las columnas 'Nombre de usuario' y 'Email'. Por favor, use la plantilla.")

        idx_username = header.index('nombre de usuario')
        idx_name = header.index('nombre') if 'nombre' in header else -1
        idx_lastname = header.index('apellidos') if 'apellidos' in header else -1
        idx_email = header.index('email')
        idx_password = header.index('contraseña') if 'contraseña' in header else -1
        idx_role = header.index('rol') if 'rol' in header else -1

        all_usernames_in_sheet = set()
        all_emails_in_sheet = set()
        to_create_data = []

        # 1. Pase de validación
        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not any(row):
                continue

            username = str(row[idx_username] or "").strip() if 0 <= idx_username < len(row) else ""
            if not username:
                raise Exception(f"Fila {row_num}: El 'Nombre de usuario' está vacío.")
            
            if username in all_usernames_in_sheet:
                raise Exception(f"Fila {row_num}: El nombre de usuario '{username}' está duplicado en el Excel.")
            all_usernames_in_sheet.add(username)

            email = str(row[idx_email] or "").strip().lower() if 0 <= idx_email < len(row) else ""
            if not email or not ExcelImporterService._is_valid_email(email):
                raise Exception(f"Fila {row_num}: El email '{email}' es inválido o está vacío.")
            
            if email in all_emails_in_sheet:
                raise Exception(f"Fila {row_num}: El email '{email}' está duplicado en el Excel.")
            all_emails_in_sheet.add(email)
            
            from django.db.models import Q
            if User.objects.filter(Q(username=username) | Q(email=email)).exists():
                raise Exception(f"Fila {row_num}: El usuario '{username}' o email '{email}' ya existe en la base de datos de PaaSify.")
                
            role = str(row[idx_role] or "").strip().lower() if 0 <= idx_role < len(row) else "student"
            if role not in ['student', 'teacher', 'admin']:
                raise Exception(f"Fila {row_num}: Rol '{role}' no válido. Use 'student', 'teacher', o 'admin'.")

            raw_pass = str(row[idx_password] or "").strip() if 0 <= idx_password < len(row) else ""
            
            to_create_data.append({
                "username": username,
                "name": str(row[idx_name] or "").strip() if 0 <= idx_name < len(row) else "",
                "lastname": str(row[idx_lastname] or "").strip() if 0 <= idx_lastname < len(row) else "",
                "email": email,
                "password": raw_pass if raw_pass else email,
                "must_change": not bool(raw_pass),  # True si se autogeneró
                "role": role
            })

        if not to_create_data:
            raise Exception("El archivo Excel está vacío o no contiene filas válidas.")

        from paasify.roles import ensure_user_group, TEACHER_GROUP_NAMES, DEFAULT_TEACHER_GROUP, STUDENT_GROUP_NAMES, DEFAULT_STUDENT_GROUP

        # 2. Transacción de Guardado
        with transaction.atomic():
            for data in to_create_data:
                user = User.objects.create_user(
                    username=data['username'],
                    email=data['email'],
                    password=data['password'],
                    first_name=data['name'],
                    last_name=data['lastname']
                )
                
                role = data['role']
                if role == 'admin':
                    user.is_staff = True
                    user.is_superuser = True
                    user.save()
                elif role == 'teacher':
                    ensure_user_group(user, TEACHER_GROUP_NAMES, DEFAULT_TEACHER_GROUP)
                elif role == 'student':
                    ensure_user_group(user, STUDENT_GROUP_NAMES, DEFAULT_STUDENT_GROUP)

                # Siempre creamos/actualizamos el UserProfile a expensas del rol (evitar IntegrityError)
                UserProfile.objects.update_or_create(
                    user=user, 
                    defaults={
                        'nombre': f"{data['name']} {data['lastname']}".strip() or data['username'],
                        'year': data['email'],
                        'must_change_password': data['must_change']
                    }
                )

        return {"success": True, "count": len(to_create_data)}
